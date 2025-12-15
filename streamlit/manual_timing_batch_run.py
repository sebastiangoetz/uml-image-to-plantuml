import os
import shutil
import sys
from pathlib import Path
from natsort import os_sorted
from openpyxl import load_workbook

import pipeline
import config
from data.multi_writer import MultiWriter

input_dir = Path("Path1")
file_list = [f for f in input_dir.glob("*") if f.is_file()]
file_list = os_sorted(file_list)
model_size = "n"
skip_first_images = 8

excel_file_path = "../thesis/evaluation/diagrams/printed/printed.xlsx"
excel_sheet_name = "Nano"

def write_list_to_row(file_path, sheet_name, row, values, first_column=1, check_name=True, expected_name=None, overwrite=False):
    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise Exception(f"Sheet '{sheet_name}' not found in the workbook.")

    ws = wb[sheet_name]

    if check_name:
        if expected_name is None:
            expected_name = values[0]
        if ws.cell(row=row, column=1).value != expected_name:
            raise Exception(f"Value in first Cell of Row {row} is '{ws.cell(row=row, column=1).value}' which is not equal to the expected diagram name '{expected_name}'.")
    else:
        if not overwrite:
            # Check if cell A in the specified row is empty
            if ws.cell(row=row, column=first_column).value is not None:
                raise Exception(f"Row {row} is not empty in column A. Aborting to prevent overwrite.")

    # Write values into the row (starting from column A)
    for col_index, value in enumerate(values, start=0):
        if value is None:
            continue
        ws.cell(row=row, column=first_column + col_index, value=value)

    # Save the file
    wb.save(file_path)
    print(f"Data written to row {row}")

for i, file in enumerate(file_list):
    file_name = file.name
    file_name_without_extension = file.stem

    print(f"Started handling of file {i + 1}/{len(file_list)}: {file_name}")

    if i < skip_first_images:
        print(f"Skipping image {file_name}")
        continue

    evaluation_data_row = []

    for run_index in range(0, 5):
        print(f"Run {run_index + 1}/5")
        processed_dir = Path(config.PROCESSED_DIR)
        if os.path.exists(processed_dir):
            shutil.rmtree(processed_dir)
        os.makedirs(processed_dir, exist_ok=True)

        result = pipeline.run_uml_extraction_pipeline(str(file), model_size=model_size, skip_diagram_rendering=True)

        evaluation_data_row.append(result["time"])

    write_list_to_row(excel_file_path, excel_sheet_name, i + 2, evaluation_data_row, first_column=125, check_name=True, expected_name=file_name_without_extension, overwrite=False)

