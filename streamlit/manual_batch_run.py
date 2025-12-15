import os
import shutil
import sys
from pathlib import Path
from natsort import os_sorted
from openpyxl import load_workbook

import pipeline
import config
from data.multi_writer import MultiWriter

input_dir = Path("Path1")
file_list = [f for f in input_dir.glob("*") if f.is_file()]
file_list = os_sorted(file_list)
skip_existing = True
model_size = "n"

excel_file_path = "../thesis/evaluation/diagrams/handwritten/handwritten.xlsx"
excel_sheet_name = "Nano"

def print_evaluations_excel_shortcut(diagram_name,
                                     number_of_detected_classes,
                                     number_of_detected_relationships,
                                     number_of_detected_class_texts,
                                     number_of_detected_labels,
                                     number_of_detected_multiplicities):
    print()
    print("Evaluation excel shortcut (only assumptions, to make counting easier, has to be checked!):")
    s = diagram_name  # Name
    s += "\t"  # Expected Classes
    s += "\t" + str(number_of_detected_classes)  # True Positive Classes
    s += "\t"  # Partial True Positive Classes
    s += "\t"  # False Positive Classes
    s += "\t"  # Total Relationships
    s += "\t"  # Expected Relationships
    s += "\t" + str(number_of_detected_relationships)  # True Positive Relationships
    s += "\t"  # False Positive Relationships
    s += "\t"  # Correct Relationship Types
    s += "\t"  # Total Class Texts
    s += "\t"  # Expected Class Texts
    s += "\t" + str(number_of_detected_class_texts)  # True Positive Class Texts
    s += "\t"  # Partial True Positive Class Texts
    s += "\t"  # False Positive Class Texts
    s += "\t"  # False Positive Class Texts in False Positive Classes
    s += "\t"  # Correct Class Text Types
    s += "\t"  # Total Labels
    s += "\t"  # Expected Labels
    s += "\t" + str(number_of_detected_labels)  # True Positive Labels
    s += "\t"  # Partial True Positive Labels
    s += "\t"  # False Positive Labels
    s += "\t"  # Correctly Assigned Labels
    s += "\t"  # Total Multiplicities
    s += "\t"  # Expected Multiplicities
    s += "\t" + str(number_of_detected_multiplicities)  # True Positive Multiplicities
    s += "\t"  # Partial True Positive Multiplicities
    s += "\t"  # False Positive Multiplicities
    s += "\t"  # Correctly Assigned Multiplicities
    s += "\t"  # Correct Multiplicities
    print(s)

def write_list_to_row(file_path, sheet_name, row, values, check_name=True, overwrite=False):
    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise Exception(f"Sheet '{sheet_name}' not found in the workbook.")

    ws = wb[sheet_name]

    if check_name:
        if ws.cell(row=row, column=1).value != values[0]:
            raise Exception(f"Value in first Cell of Row {row} is '{ws.cell(row=row, column=1).value}' which is not equal to the expected diagram name '{values[0]}'.")
    else:
        if not overwrite:
            # Check if cell A in the specified row is empty
            if ws.cell(row=row, column=1).value is not None:
                raise Exception(f"Row {row} is not empty in column A. Aborting to prevent overwrite.")

    # Write values into the row (starting from column A)
    for col_index, value in enumerate(values, start=1):
        if value is None:
            continue
        ws.cell(row=row, column=col_index, value=value)

    # Save the file
    wb.save(file_path)
    print(f"Data written to row {row}")

original_stdout = sys.stdout  # save original console

for i, file in enumerate(file_list):
    file_name = file.name
    file_name_without_extension = file.stem

    print(f"Started handling of file {i + 1}/{len(file_list)}: {file_name}")

    config.PROCESSED_DIR = str(input_dir) + "/" +  file_name_without_extension

    processed_dir = Path(config.PROCESSED_DIR)
    if os.path.exists(processed_dir):
        if skip_existing:
            continue
        shutil.rmtree(processed_dir)
    os.makedirs(processed_dir, exist_ok=True)

    # Redirect to both console and log file
    log_file_path = Path(config.PROCESSED_DIR + "/" + "log.txt")
    with open(log_file_path, "w", encoding="utf-8") as log_file:
        sys.stdout = MultiWriter(original_stdout, log_file)

        result = pipeline.run_uml_extraction_pipeline(str(file), model_size=model_size, skip_diagram_rendering=True)

        evaluation_data_row = []
        evaluation_data_row.append(result["diagram_name"])
        evaluation_data_row.append(None)
        evaluation_data_row.append(result["number_of_detected_classes"])
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(result["number_of_detected_relationships"])
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(result["number_of_detected_class_texts"])
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(result["number_of_detected_labels"])
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(result["number_of_detected_multiplicities"])
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)
        evaluation_data_row.append(None)

        write_list_to_row(excel_file_path, excel_sheet_name, i + 2, evaluation_data_row, check_name=True, overwrite=False)

    # Restore stdout to console
    sys.stdout = original_stdout
